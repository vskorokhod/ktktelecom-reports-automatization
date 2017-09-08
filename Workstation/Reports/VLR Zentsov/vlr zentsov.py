import subprocess
from datetime import date, timedelta, datetime
from openpyxl import load_workbook


def _transfer_data(file_from, dir_to):
    wb_from = load_workbook(file_from)
    ws_from = wb_from.worksheets[0]
    
    week_ago = datetime.strftime(date.today() - timedelta(days=7), '%d.%m.%Y')
    today = datetime.strftime(date.today(), '%d.%m.%Y')
    
    wb_to = load_workbook('{}VLR {}.xlsx'.format(dir_to, week_ago))
    ws_to = wb_to.worksheets[0]
    
    for i in range(1, ws_from.max_row + 1):
        for j in range(1, ws_from.max_column + 1):
            ws_to.cell(row=i + 1, column=j).value = ws_from.cell(row=i, column=j).value
            
    wb_to.save('{}VLR {}.xlsx'.format(dir_to, today))


def _report_composition():    
    commands = ['"C:\Program Files (x86)\WinSCP\winscp.com" /ini=nul /command "open sftp://ossuser:Changeme_123@10.10.180.65 -hostkey=""ssh-rsa 2048 61:5b:c1:1f:ef:64:75:52:53:11:b1:d5:3f:6b:ae:4b""" "get -delete ""/opt/oss/server/var/fileint/pm/NOC_reports/VLR_Zentsov/*.xlsx"" ""C:\Reports\VLR Zentsov\\""" "close" "exit"',
                'del "C:\Reports\VLR Zentsov\*.xlsx"']
     
    subprocess.run(commands[0], shell=True)
    _transfer_data("C:\Reports\VLR Zentsov\VLR Subscribers.xlsx", "Z:\ТД\Мониторинг\Отчёты\Недельный отчёт VLR Зенцову - 9ч. пятница\\")
    subprocess.run(commands[1], shell=True)

        
if __name__ == '__main__':
    _report_composition()