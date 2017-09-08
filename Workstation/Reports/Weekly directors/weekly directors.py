import subprocess
from openpyxl import load_workbook


def _transfer_data(file_from, dir_to):
    subfolders = ['CSSR_15\\', 
                  'Call Drop rate 2G_15\\', 
                  'Call Drop rate 3G_15\\', 
                  'Voice Trafic 2G 3G_60\\', 
                  'Data Traffic 2G GBytes_60\\', 
                  'Data Traffic 3G GBytes_60\\', 
                  'Data Traffic 4G GBytes_60\\', 
                  '2G Roamers Data Traffic_60\\', 
                  '3G Roamers Data Traffic_60\\', 
                  '4G Roamers Data Traffic_60\\']
    files_to = ['{}{}11.xlsx'.format(dir_to, subfolder) for subfolder in subfolders]
    
    wb_from = load_workbook(file_from)
    for file_to in files_to:
        wb_to = load_workbook(file_to)
        ws_from = wb_from.worksheets[files_to.index(file_to)]
        ws_to = wb_to.worksheets[0]
        for i in range(1, ws_from.max_row + 1):
            for j in range(1, ws_from.max_column + 1):
                ws_to.cell(row=i + 8, column=j).value = ws_from.cell(row=i, column=j).value
        wb_to.save(file_to)


def _report_composition():    
    commands = ['"C:\Program Files (x86)\WinSCP\winscp.com" /ini=nul /command "open sftp://ossuser:Changeme_123@10.10.180.65 -hostkey=""ssh-rsa 2048 61:5b:c1:1f:ef:64:75:52:53:11:b1:d5:3f:6b:ae:4b""" "get -delete ""/opt/oss/server/var/fileint/pm/NOC_reports/Weekly_directors/*.xlsx"" ""C:\Reports\Weekly directors\\""" "close" "exit"',
                'del "C:\Reports\Weekly directors\*.xlsx"']
     
    subprocess.run(commands[0], shell=True)
    _transfer_data("C:\Reports\Weekly directors\Weekly directors report.xlsx", "Z:\ТД\Мониторинг\Отчёты\Недельный отчёт директорам-к 11ч. пятницы\Report\\")
    subprocess.run(commands[1], shell=True)

        
if __name__ == '__main__':
    _report_composition()