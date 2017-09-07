import subprocess
from openpyxl import load_workbook


def _transfer_data(file_from, files_to):
    wb_from = load_workbook(file_from)
    for file_to in files_to:
        wb_to = load_workbook(file_to)
        for i in range(4):
            ws_from = wb_from.worksheets[i + 4 * files_to.index(file_to)]
            ws_to = wb_to.worksheets[i]
            for row in ws_from.rows:
                ws_to.append([item.value for item in row])
        wb_to.save(file_to)


def _report_composition():    
    commands = ['"C:\Program Files (x86)\WinSCP\winscp.com" /ini=nul /command "open sftp://ossuser:Changeme_123@10.10.180.65 -hostkey=""ssh-rsa 2048 61:5b:c1:1f:ef:64:75:52:53:11:b1:d5:3f:6b:ae:4b""" "get -delete ""/opt/oss/server/var/fileint/pm/NOC_reports/Traffic_utilization/*.xlsx"" ""C:\Reports\Traffic utilization\\""" "close" "exit"',
                'del "C:\Reports\Traffic utilization\*.xlsx"']
     
    subprocess.run(commands[0], shell=True)
    _transfer_data("C:\Reports\Traffic utilization\Traffic utilization report.xlsx", ["G:\Share\ТД\Мониторинг\Отчёты\Недельный отчёт ТГ Гончаруку - пятница\Report Traffic Trunk Groups ISUP.xlsx", "G:\Share\ТД\Мониторинг\Отчёты\Недельный отчёт ТГ Гончаруку - пятница\Report Traffic Trunk Groups SIP.xlsx"])
    subprocess.run(commands[1], shell=True)

        
if __name__ == '__main__':
    _report_composition()