import os
import re
import csv
from datetime import date, timedelta, datetime
from openpyxl import Workbook, load_workbook


def _get_files_by_path(path, file_id):
    return ['{0}{1}'.format(path, filename) for filename in os.listdir(path) if file_id in filename]

    
def _compose_files_list(file_id, daily_report=False, kpi_weekly_report=False):
    yesterday = datetime.strftime(date.today() - timedelta(days=1), '%Y%m%d')
    path = '../../pmexport_{0}/'.format(yesterday)
    files = _get_files_by_path(path, file_id)
    if daily_report:
        files = files[-3:]
        today = datetime.strftime(date.today(), '%Y%m%d')
        path = '../../pmexport_{0}/'.format(today)
        files.extend(_get_files_by_path(path, file_id))
    return files


def _get_sheet_name(file_id):
    sheet_name = file_id
    replacement_map = zip(['Trunk Groups', 'Office Directions', 'Congection Rate_60'], ['ISUP', 'SIP', 'Trunk Utilization_OD'])
    for left, right in replacement_map:
        sheet_name = sheet_name.replace(left, right)
    sheet_name = re.sub('(\d\.\d )|(\d+_)|(.csv)', '', sheet_name)
    return sheet_name
    
    
def _remove_old_data(wb, ws, daily_report=False, kpi_weekly_report=False):
    if ws.title in ['HSS SPS MSS USN CPU Usage', 'UMG CPU Usage', 'UGW CPU Usage', 'M3UA MSC Signaling Links']:
        ws._current_row = 0
        return
    
    if daily_report:
        week_ago = datetime.strftime(date.today() - timedelta(days=7), '%d.%m.%Y')
        sought_str = '{0} 21:00:00'.format(week_ago)
    elif kpi_weekly_report:
        prev_month_length = {'01': 31, '02': 31, '03': 28, '04': 31, '05': 30, '06': 31, '07': 30, '08': 31, '09': 31, '10': 30, '11': 31, '12': 30}
        delta = prev_month_length[datetime.strftime(date.today(), '%m')]
        month_ago = datetime.strftime(date.today() - timedelta(days=delta), '%d.%m.%Y')
        sought_str = '{0} 00:00:00'.format(month_ago)
    
    for cell in ws['A']:
        if cell.value == sought_str:
            offset = cell.row - 1
            break
            
    for i in range(1, ws.max_row - offset + 1):
        for j in range(1, ws.max_column + 1):
            ws.cell(row=i, column=j).value = ws.cell(row=i+offset, column=j).value
            
    ws._current_row = ws.max_row - offset
   
   
def _open_workbook(file_id, output_file, daily_report=False, kpi_weekly_report=False):
    sheet_name = _get_sheet_name(file_id)
    try:
        wb = load_workbook(output_file)
    except IOError:
        wb = Workbook(output_file)
    try:
        ws = wb[sheet_name]
    except KeyError:
        ws = wb.create_sheet(sheet_name)
        
    if (daily_report or kpi_weekly_report):
        _remove_old_data(wb, ws, daily_report, kpi_weekly_report)
        
    return wb, ws

    
def _normalize_row(row, permutation_level):
    YEAR, MONTH, DAY, TIME = slice(0, 4), slice(5, 7), slice(8, 10), slice(11, 16)
    row[0] = '{0}.{1}.{2} {3}:00'.format(row[0][DAY], row[0][MONTH], row[0][YEAR], row[0][TIME])
    row[2], row[3] = row[2].split('/')
    row[3] = row[3][row[3].index(':') + 1:]
    if permutation_level == 1:
        row[4], row[5] = row[5], row[4]
    elif permutation_level == 2:
        row[4], row[5], row[6], row[7] = row[7], row[6], row[4], row[5]
	
    
def _append_to_workbook(ws, files, file_id):
    for filename in files:        
        with open(filename) as f:
            f.readline()
            f.readline()
            for line in csv.reader(f):
                current_row = line
                if any(expr in file_id for expr in ['CSSR.csv', 'PDP SR', 'M3UA MSC']):
                    permutation_level = 1
                elif 'VLR_Subscribers' in file_id:
                    permutation_level = 2
                else:
                    permutation_level = None
                _normalize_row(current_row, permutation_level)
                ws.append(current_row)
    return current_row


def _set_numeric_types(ws, row):
    for col in range(0, ws.max_column):
        try:
            float(row[col])
            for cell in ws[chr(65 + col)]:
                cell.data_type = 'n'
        except ValueError:
            continue

    
def collect_data(file_id, output_file, **kwargs):   
    files = _compose_files_list(file_id, **kwargs)
    wb, ws = _open_workbook(file_id, output_file, **kwargs)
    row = _append_to_workbook(ws, files, file_id)
    _set_numeric_types(ws, row)
    wb.save(output_file)